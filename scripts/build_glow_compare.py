# massa 가격표에 GLOW 하노이 시세를 나란히 붙여 비교 엑셀을 만든다.
import csv, json, re, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "massa_GLOW_가격비교.xlsx"
INK, AMBER, SAND, LINE = "2A2320", "B0742F", "F1E9DE", "E2D8CB"
LOW, HIGH, INPUT = "E7F1EA", "FBE7EC", "FFFF00"     # 싸다 / 비싸다 / 입력칸

F = lambda **k: Font(name="Arial", **k)
thin = Side(style="thin", color=LINE)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

rows = [r for r in csv.reader(open("services.csv", encoding="utf-8-sig")) if r and len(r) == 6]
base = lambda n: re.sub(r"\s*\d+\s*분\s*$", "", n).strip()
G = json.load(open("glow_prices.json", encoding="utf-8"))
glow = {(r["name"], r["dur"]): r for r in G["rows"]}

# massa 서비스 → GLOW 대응. 시간이 다르면 세 번째 값에 GLOW 쪽 시간을 적는다.
MAP = {
    "아로마테라피": ("Aroma Massage", None), "아로마 마사지": ("Aroma Massage", None),
    "스웨디시": ("Oil Massage", None),
    "오일 마사지 + 부항 요법": ("Oil Massage", None),
    "오일 없는 마사지": ("Massage Without Oil", None),
    "타이 마사지": ("Thai Massage", None), "태국식 마사지": ("Thai Massage", None),
    "핫 스톤마사지": ("Hot Stone Massage", None),
    "어깨·목 마사지": ("Neck Shoulder Massage", None),
    "목·어깨 테라피": ("Neck Shoulder Massage", None),
    "등 테라피": ("Back Therapy", None),
    "다리 마사지": ("Foot Massage", None), "풋 테라피": ("Foot Massage", None),
    "머리 마사지": ("Head Therapy", None), "헤드 테라피": ("Head Therapy", None),
    "스포츠 테라피": ("Sports Therapy", None),
    "귀청소": ("Earwax", 40),
    "한국식 전신 스크럽": ("Korean Full Body Scrub", 60),
    "비키니 왁싱": ("Bikini Waxing", 60),
    "겨드랑이 왁싱": ("Underarm Waxing", 60),
}
CAT_KO = {"massage": "마사지", "therapist_care": "홈뷰티·케어"}

wb = Workbook()

def head(ws, labels, widths, row):
    for i, (t, w) in enumerate(zip(labels, widths), 1):
        c = ws.cell(row, i, t)
        c.font = F(bold=True, size=10, color=INK)
        c.fill = PatternFill("solid", fgColor=SAND)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 34
    ws.freeze_panes = ws.cell(row + 1, 1)

# ── 1. 비교표 ────────────────────────────────────────────────
ws = wb.active
ws.title = "GLOW 비교"
ws["A1"] = "massa 가격 vs GLOW 하노이 시세"
ws["A1"].font = F(bold=True, size=15, color=INK)
ws["A2"] = (f"GLOW 는 기사가 각자 값을 정합니다. 아래 GLOW 값은 하노이 기사 {G['표본수']}명을 모아 낸 "
            "최저·중앙·최고입니다. 초록은 massa 가 더 싸다, 분홍은 더 비싸다는 뜻입니다. "
            "노란 칸(K열)에 새 가격을 넣으면 GLOW 중앙값 대비 위치가 다시 계산됩니다.")
ws["A2"].font = F(size=9, color="6B5F57")
ws.merge_cells("A2:M2")
ws["A3"] = "출처: " + G["출처"]
ws["A3"].font = F(size=9, color="6B5F57")

HDR = ["카테고리", "massa 서비스", "시간(분)", "massa 가격", "GLOW 대응 서비스",
       "GLOW 시간", "GLOW 최저", "GLOW 중앙", "GLOW 최고", "중앙 대비",
       "새 가격 (입력)", "새 가격의 중앙 대비", "표본"]
head(ws, HDR, [11, 24, 8, 13, 22, 9, 12, 12, 12, 11, 13, 13, 7], row=5)

r = 6
for cat, mt, name, dur, price, act in sorted(rows, key=lambda x: (x[0], base(x[2]), int(x[3]))):
    b, dur, price = base(name), int(dur), int(price)
    gname, gdur = MAP.get(b, (None, None))
    g = glow.get((gname, gdur or dur)) if gname else None

    ws.cell(r, 1, CAT_KO.get(cat, cat)).font = F(size=10)
    ws.cell(r, 2, name).font = F(size=10)
    ws.cell(r, 3, dur).font = F(size=10)
    ws.cell(r, 4, price).font = F(size=10)
    ws.cell(r, 5, gname or "대응 없음").font = F(size=10, color=INK if g else "9A8F84")
    if g:
        ws.cell(r, 6, g["dur"]).font = F(size=10)
        ws.cell(r, 7, g["min"]).font = F(size=10)
        ws.cell(r, 8, g["med"]).font = F(bold=True, size=10)
        ws.cell(r, 9, g["max"]).font = F(size=10)
        ws.cell(r, 10, f"=IFERROR(D{r}/H{r}-1,\"\")").font = F(size=10, bold=True)
        ws.cell(r, 12, f'=IF(K{r}="","",IFERROR(K{r}/H{r}-1,""))').font = F(size=10)
        ws.cell(r, 13, g["n"]).font = F(size=9, color="6B5F57")
        # massa 가 더 싼가 비싼가를 색으로
        tone = LOW if price < g["med"] else (HIGH if price > g["med"] else None)
        if tone:
            for c in (4, 8, 10):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=tone)
        if gdur and gdur != dur:
            ws.cell(r, 6).font = F(size=10, bold=True, color=AMBER)
    ws.cell(r, 11).fill = PatternFill("solid", fgColor=INPUT)
    ws.cell(r, 11).font = F(size=10, color="0000FF")

    for c in (4, 7, 8, 9, 11):
        ws.cell(r, c).number_format = "#,##0"
    for c in (10, 12):
        ws.cell(r, c).number_format = "+0.0%;[Red]-0.0%;0.0%"
    for c in (3, 6, 13):
        ws.cell(r, c).alignment = Alignment(horizontal="center")
    for c in range(1, 14):
        ws.cell(r, c).border = BOX
    r += 1

end = r - 1
ws.cell(r + 1, 2, "대응이 있는 항목만 평균").font = F(bold=True, size=10)
ws.cell(r + 1, 10, f'=IFERROR(AVERAGE(J6:J{end}),"")').font = F(bold=True, size=10)
ws.cell(r + 1, 10).number_format = "+0.0%;[Red]-0.0%;0.0%"
ws.cell(r + 2, 2, "GLOW 대응이 없는 항목 수").font = F(size=10)
ws.cell(r + 2, 4, f'=COUNTIF(E6:E{end},"대응 없음")').font = F(size=10)

# ── 2. 시간대별 나란히 ───────────────────────────────────────
ws2 = wb.create_sheet("시간대별 나란히")
ws2["A1"] = "60 · 90 · 120분 — massa 와 GLOW 중앙값"
ws2["A1"].font = F(bold=True, size=15, color=INK)
ws2["A2"] = "GLOW 는 시간이 늘 때마다 10만동씩 일정하게 올립니다. massa 와 붙는 지점이 어디인지 보십시오."
ws2["A2"].font = F(size=9, color="6B5F57")
ws2.merge_cells("A2:J2")

head(ws2, ["카테고리", "massa 서비스", "60분", "90분", "120분",
           "GLOW 60분", "GLOW 90분", "GLOW 120분", "60분 차이", "120분 차이"],
     [11, 24, 12, 12, 12, 12, 12, 12, 11, 11], row=4)

grid = collections.defaultdict(dict)
for cat, mt, name, dur, price, act in rows:
    grid[(cat, base(name))][int(dur)] = int(price)

r = 5
for (cat, b) in sorted(grid, key=lambda k: (k[0], k[1])):
    d = grid[(cat, b)]
    gname, gdur = MAP.get(b, (None, None))
    ws2.cell(r, 1, CAT_KO.get(cat, cat)).font = F(size=10)
    ws2.cell(r, 2, b).font = F(size=10)
    for i, du in enumerate([60, 90, 120], 3):
        ws2.cell(r, i, d.get(du)).font = F(size=10)
        ws2.cell(r, i).number_format = "#,##0"
        g = glow.get((gname, du)) if gname and not gdur else None
        ws2.cell(r, i + 3, g["med"] if g else None).font = F(size=10, color="6B5F57")
        ws2.cell(r, i + 3).number_format = "#,##0"
    ws2.cell(r, 9,  f'=IF(OR(C{r}="",F{r}=""),"",C{r}/F{r}-1)').font = F(size=10, bold=True)
    ws2.cell(r, 10, f'=IF(OR(E{r}="",H{r}=""),"",E{r}/H{r}-1)').font = F(size=10, bold=True)
    for c in (9, 10):
        ws2.cell(r, c).number_format = "+0.0%;[Red]-0.0%;0.0%"
    for c in range(1, 11):
        ws2.cell(r, c).border = BOX
    r += 1

# ── 3. 읽을 거리 ─────────────────────────────────────────────
ws3 = wb.create_sheet("무엇이 보이나")
ws3["A1"] = "숫자에서 읽히는 것"
ws3["A1"].font = F(bold=True, size=15, color=INK)
ws3["A2"] = "판단은 사장님이 하십니다. 아래는 표를 요약한 것입니다."
ws3["A2"].font = F(size=9, color="6B5F57")
head(ws3, ["항목", "무엇이", "숫자", "생각해 볼 것"], [16, 30, 40, 42], row=4)

d = G["대표가_분포"]
NOTES = [
    ("시장 표준", "GLOW 하노이의 사실상 정가",
     f"60분: 500,000₫ {d['60분']['500000']}명 · 600,000₫ {d['60분']['600000']}명 · "
     f"700,000₫ {d['60분']['700000']}명 (총 {G['표본수']}명). "
     "GLOW 는 시간이 늘 때마다 10만동씩 일정하게 올립니다.",
     "massa 도 '60분 500,000 · 90분 600,000 · 120분 700,000' 처럼 규칙을 하나로 정할지."),
    ("massa 가 싸다", "60분 마사지 대부분",
     "massa 350,000~550,000 대 GLOW 중앙값 500,000. "
     "다리·머리·풋 350,000 은 GLOW 대비 -30% 입니다.",
     "싸게 들어가 점유율을 살지, 시세에 붙여 마진을 살지."),
    ("massa 가 비싸다", "120분 전 항목",
     "massa 대부분 850,000 · 핫스톤/아로마테라피 950,000 대 GLOW 중앙값 700,000. "
     "+21% ~ +36% 입니다.",
     "가장 눈에 띄는 역전입니다. 120분을 내릴지, 프리미엄으로 남길지."),
    ("이상한 지점", "시간이 길수록 분당 단가가 오른다",
     "massa 다리 마사지 60분 5,833₫/분 → 120분 7,083₫/분. "
     "GLOW 는 8,333 → 5,833 으로 내려갑니다.",
     "보통 장시간은 할인합니다. 지금 구조는 반대입니다."),
    ("귀청소", "massa 30분 200,000 대 GLOW 40분 400,000~500,000",
     "분당으로 봐도 massa 6,667₫ 대 GLOW 10,000₫. 절반 수준입니다.",
     "귀청소는 올릴 여지가 큽니다. 시간도 40분으로 맞출지."),
    ("한국식 스크럽", "massa 60분 500,000 대 GLOW 60분 600,000~700,000",
     "-17% ~ -29%. 한국식은 massa 의 강점인데 값이 더 쌉니다.",
     "간판 상품이라면 시세만큼은 받을지."),
    ("왁싱", "비키니 massa 45분 500,000 대 GLOW 60분 600,000~700,000",
     "겨드랑이는 massa 30분 250,000 대 GLOW 60분 450,000. 시간이 달라 직접 비교는 조심해야 합니다.",
     "왁싱 시간 단위를 GLOW 와 맞출지."),
    ("대응 없음", "네일·젤·각질·부위별 왁싱 일부",
     "GLOW 하노이 프로필에서 같은 항목을 찾지 못했습니다. "
     "GLOW 앱 안에는 있을 수 있으나 공개 페이지에는 없습니다.",
     "경쟁이 적은 영역일 수 있습니다. 값을 스스로 정해도 됩니다."),
    ("표본의 한계", "이 값은 공개 프로필 기준",
     f"하노이 기사 {G['표본수']}명의 공개 페이지에서 긁었습니다. "
     "실제 앱 안 가격·할인·프로모션은 다를 수 있습니다.",
     "중요한 결정 전에는 GLOW 앱에서 직접 한 번 확인하실 것."),
]
r = 5
for a, b_, c_, d_ in NOTES:
    ws3.cell(r, 1, a).font = F(size=10, bold=True, color=AMBER)
    ws3.cell(r, 2, b_).font = F(size=10)
    ws3.cell(r, 3, c_).font = F(size=10)
    ws3.cell(r, 4, d_).font = F(size=10)
    for c in range(1, 5):
        ws3.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.cell(r, c).border = BOX
    ws3.row_dimensions[r].height = 54
    r += 1

wb.save(OUT)
print("저장:", OUT, "· massa", len(rows), "항목 · GLOW 표본", G["표본수"], "명")
