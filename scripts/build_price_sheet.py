# services 테이블을 가격 재정비용 엑셀로 만든다. 중복·편차를 눈에 띄게 표시하는 게 목적이다.
import csv, re, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "services.csv"
OUT = "massa_가격표.xlsx"

INK   = "2A2320"
AMBER = "B0742F"
SAND  = "F1E9DE"
LINE  = "E2D8CB"
WARN  = "FDF0D5"
INPUT = "FFFF00"

F  = lambda **k: Font(name="Arial", **k)
thin = Side(style="thin", color=LINE)
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)

rows = [r for r in csv.reader(open(SRC, encoding="utf-8-sig")) if r and len(r) == 6]
base = lambda n: re.sub(r"\s*\d+\s*분\s*$", "", n).strip()

grid = collections.defaultdict(dict)          # (카테고리, 기본명) -> {시간: 가격}
for cat, mt, name, dur, price, act in rows:
    grid[(cat, base(name))][int(dur)] = int(price)

# 시간대별 가격이 완전히 같은 기본명끼리 묶으면 그게 곧 이름만 다른 중복이다
same = collections.defaultdict(list)
for (cat, b), d in grid.items():
    same[(cat, tuple(sorted(d.items())))].append(b)
dups = {}
for (cat, _), names in same.items():
    if len(names) > 1:
        for n in names:
            dups[(cat, n)] = " · ".join(sorted(x for x in names if x != n))

CAT_KO = {"massage": "마사지", "therapist_care": "홈뷰티·케어"}
wb = Workbook()

def head(ws, labels, widths, row=1):
    for i, (t, w) in enumerate(zip(labels, widths), 1):
        c = ws.cell(row, i, t)
        c.font = F(bold=True, size=10, color=INK)
        c.fill = PatternFill("solid", fgColor=SAND)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row + 1, 1)

# ── 1. 현재 가격표 ────────────────────────────────────────────
ws = wb.active
ws.title = "현재 가격표"
ws["A1"] = "massa 서비스 가격 현황"
ws["A1"].font = F(bold=True, size=15, color=INK)
ws["A2"] = ("노란 칸(F열)에 새 가격을 넣으면 변동률이 자동 계산됩니다. "
            "맨 오른쪽 열은 시간대별 가격이 완전히 똑같은 항목을 짝지어 보여 줍니다 — "
            "이름만 다른 중복일 수도, 값이 우연히 같은 다른 시술일 수도 있습니다. 판단은 ‘정리 필요’ 탭에 적어 두었습니다.")
ws["A2"].font = F(size=9, color="6B5F57")
ws.merge_cells("A2:I2")

HDR = ["카테고리", "서비스", "시간(분)", "현재가 (VND)", "분당 단가",
       "새 가격 (입력)", "변동률", "활성", "가격이 같은 항목"]
head(ws, HDR, [12, 26, 9, 14, 11, 14, 9, 7, 30], row=4)

r = 5
for cat, mt, name, dur, price, act in sorted(
        rows, key=lambda x: (x[0], base(x[2]), int(x[3]))):
    b = base(name)
    ws.cell(r, 1, CAT_KO.get(cat, cat)).font = F(size=10)
    ws.cell(r, 2, name).font = F(size=10)
    ws.cell(r, 3, int(dur)).font = F(size=10)
    ws.cell(r, 4, int(price)).font = F(size=10)
    ws.cell(r, 5, f"=IFERROR(D{r}/C{r},\"\")").font = F(size=10, color="6B5F57")
    ws.cell(r, 6).fill = PatternFill("solid", fgColor=INPUT)
    ws.cell(r, 6).font = F(size=10, color="0000FF")
    ws.cell(r, 7, f'=IF(F{r}="","",IFERROR(F{r}/D{r}-1,""))').font = F(size=10)
    ws.cell(r, 8, act).font = F(size=10)
    d = dups.get((cat, b))
    ws.cell(r, 9, ("동일: " + d) if d else "").font = F(size=9, color=AMBER)

    ws.cell(r, 4).number_format = "#,##0"
    ws.cell(r, 5).number_format = "#,##0"
    ws.cell(r, 6).number_format = "#,##0"
    ws.cell(r, 7).number_format = "0.0%;[Red]-0.0%;-"
    ws.cell(r, 3).alignment = Alignment(horizontal="center")
    ws.cell(r, 8).alignment = Alignment(horizontal="center")
    for c in range(1, 10):
        ws.cell(r, c).border = BOX
        # 중복인 줄은 배경으로 표시한다. 입력 칸(F)만은 노란색을 지키게 건너뛴다
        if d and c != 6:
            ws.cell(r, c).fill = PatternFill("solid", fgColor=WARN)
    r += 1

tot = r
ws.cell(tot, 2, f"합계 {len(rows)}개 항목").font = F(bold=True, size=10)
ws.cell(tot, 4, f"=SUM(D5:D{r-1})").font = F(bold=True, size=10)
ws.cell(tot, 4).number_format = "#,##0"
ws.cell(tot, 6, f"=IF(COUNT(F5:F{r-1})=0,\"\",SUM(F5:F{r-1}))").font = F(bold=True, size=10)
ws.cell(tot, 6).number_format = "#,##0"
ws.cell(tot + 2, 2, "출처: massa 운영 DB public.services (2026-09-06 기준)").font = F(size=9, color="6B5F57")

# ── 2. 시간대별 비교 ──────────────────────────────────────────
ws2 = wb.create_sheet("시간대별 비교")
ws2["A1"] = "같은 서비스의 시간대별 가격"
ws2["A1"].font = F(bold=True, size=15, color=INK)
ws2["A2"] = "빈칸은 그 시간대 상품이 없다는 뜻입니다. 가로로 읽으면 시간이 늘 때 가격이 어떻게 붙는지 보입니다."
ws2["A2"].font = F(size=9, color="6B5F57")
ws2.merge_cells("A2:H2")

DURS = [30, 45, 60, 90, 120]
head(ws2, ["카테고리", "서비스"] + [f"{d}분" for d in DURS] + ["60→120 배수"],
     [12, 26] + [12] * 5 + [13], row=4)

r = 5
for (cat, b) in sorted(grid, key=lambda k: (k[0], k[1])):
    d = grid[(cat, b)]
    ws2.cell(r, 1, CAT_KO.get(cat, cat)).font = F(size=10)
    ws2.cell(r, 2, b).font = F(size=10)
    for i, du in enumerate(DURS, 3):
        c = ws2.cell(r, i, d.get(du))
        c.font = F(size=10)
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right")
    # 60분 대비 120분이 몇 배인가. 2배를 크게 벗어나면 가격표가 뒤틀려 있다는 뜻이다
    ws2.cell(r, 8, f'=IF(OR(E{r}="",G{r}=""),"",IFERROR(G{r}/E{r},""))').font = F(size=10)
    ws2.cell(r, 8).number_format = '0.00"배";;-'
    for c in range(1, 9):
        ws2.cell(r, c).border = BOX
    r += 1

# ── 3. 정리 필요 항목 ─────────────────────────────────────────
ws3 = wb.create_sheet("정리 필요")
ws3["A1"] = "가격 재정비 전에 결정해야 할 것"
ws3["A1"].font = F(bold=True, size=15, color=INK)
ws3["A2"] = "DB 를 그대로 읽어 뽑은 것입니다. 판단은 사장님이 하시고, 정하시면 반영하겠습니다."
ws3["A2"].font = F(size=9, color="6B5F57")
ws3.merge_cells("A2:D2")

head(ws3, ["구분", "무엇이", "지금 상태", "결정할 것"], [14, 30, 46, 34], row=4)

ISSUES = [
    ("이름 중복", "다리 마사지 / 풋 테라피",
     "60·90·120분 가격이 350,000 / 500,000 / 850,000 으로 완전히 같습니다.",
     "하나로 합칠지, 다른 시술로 나눠 값을 다르게 할지."),
    ("이름 중복", "머리 마사지 / 헤드 테라피",
     "60·90·120분 가격이 350,000 / 500,000 / 850,000 으로 완전히 같습니다.",
     "하나로 합칠지."),
    ("이름 중복", "목·어깨 테라피 / 어깨·목 마사지 / 등 테라피",
     "셋 다 400,000 / 650,000 / 850,000 입니다. 이름만 다릅니다.",
     "셋을 하나로 줄일지, 등과 목·어깨를 갈라 값을 다르게 할지."),
    ("이름 중복", "타이 마사지 / 태국식 마사지",
     "60·90분 값이 같습니다(400,000 / 550,000). 120분은 태국식에만 있습니다.",
     "이름 하나로 통일하고 120분을 둘지."),
    ("값 불일치", "아로마 마사지 / 아로마테라피",
     "같은 아로마인데 전 시간대에서 100,000 씩 차이 납니다 "
     "(450·650·850 대 550·750·950).",
     "등급을 나눈 것이라면 이름에 드러내고, 아니면 값을 맞출 것."),
    ("값 눌림", "120분 요금이 대부분 850,000",
     "60분에서 350,000 과 500,000 으로 벌어졌던 항목이 120분에서는 모두 "
     "850,000 으로 모입니다. 시간이 길수록 시술 간 값 차이가 사라집니다.",
     "120분에도 등급 차이를 남길지."),
    ("단가 역전", "다리·머리·풋 테라피",
     "60분 5,833원/분 대비 120분 7,083원/분 으로 길수록 분당 단가가 올라갑니다. "
     "보통은 길수록 내려갑니다.",
     "장시간 할인 구조로 갈지, 지금 구조를 유지할지."),
    ("빠진 시간", "스웨디시 · 타이 마사지",
     "다른 마사지에는 다 있는 120분이 이 둘에는 없습니다.",
     "추가할지, 일부러 뺀 것인지."),
    ("값 우연 일치", "왁싱 60분 / 한국식 전신 스크럽 60분 · 스포츠 테라피 / 오일 마사지+부항 요법",
     "서로 다른 시술인데 값이 똑같습니다(각각 500,000 · 500,000/700,000/850,000). "
     "‘가격이 같은 항목’ 열에 짝으로 뜨지만 이름 중복은 아닙니다.",
     "값을 갈라 등급을 드러낼지, 그대로 둘지."),
    ("항목 겹침", "왁싱 60분 (500,000)",
     "부위별 왁싱(겨드랑이·복부·가슴·팔·다리·비키니)이 따로 있는데 "
     "뭉뚱그린 ‘왁싱 60분’ 이 함께 있습니다.",
     "전신 패키지로 이름을 바꿀지, 없앨지."),
    ("분류 비어 있음", "massage_type 컬럼",
     "60개 중 8개에만 값이 있습니다(아로마·스웨디시·타이). 나머지 52개는 비어 있어 "
     "앱에서 종류로 거르면 걸리지 않습니다.",
     "전부 채울지, 이 분류를 안 쓸지."),
]
r = 5
for kind, what, now, todo in ISSUES:
    ws3.cell(r, 1, kind).font = F(size=10, bold=True, color=AMBER)
    ws3.cell(r, 2, what).font = F(size=10)
    ws3.cell(r, 3, now).font = F(size=10)
    ws3.cell(r, 4, todo).font = F(size=10)
    for c in range(1, 5):
        ws3.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.cell(r, c).border = BOX
    ws3.row_dimensions[r].height = 46
    r += 1

wb.save(OUT)
print("저장:", OUT, "· 항목", len(rows), "· 중복군", len({v for v in dups.values()}))
